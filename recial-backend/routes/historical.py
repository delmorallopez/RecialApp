"""
Historical mass-balance import — Level 1 (monthly summary totals).
Add this route to routes/reports.py (or a new routes/historical.py).

Reads a legacy mass-balance workbook (2020-2025), which has:
  - a "YYYY" tab: individual receipts, with entrance-lot subtotal rows
    interleaved (lot codes like 010124A in the FECHA column, blank ID).
    Receipt IDs look like 01B, 85/1A (NN + optional /N + A|B).
  - a "Mass Balance" tab: entrances block, disposals (DESECHO SOLIDO),
    dispatches (SA... lots), and an opening STOCK row.

Level 1 stores MONTHLY receipt totals + yearly dispatch/disposal/stock totals
into historical_monthly_summary / historical_year_meta. It does NOT create
live receipts/dispatches (those would be un-traceable).

Required imports at the top of the route file:
    from fastapi import UploadFile, File
    from io import BytesIO
    from openpyxl import load_workbook
    import re
    from models.historical_summary import HistoricalMonthlySummary, HistoricalYearMeta
"""

import re
from io import BytesIO
from datetime import datetime
from fastapi import UploadFile, File, APIRouter, Depends, HTTPException
from io import BytesIO
from openpyxl import load_workbook
from models.historical_summary import HistoricalMonthlySummary, HistoricalYearMeta
from sqlalchemy.orm import Session
from database import get_db
from auth import require_admin
from models.users import User

router = APIRouter(tags=["Historical"])


_RECEIPT_ID = re.compile(r'^\d+(/\d+)?[AB]$')     # 01B, 85/1A
_ENTRANCE_LOT = re.compile(r'^\d{6}[AB]$')         # 010124A
_DISPATCH_LOT = re.compile(r'^SA\d+', re.IGNORECASE)


def _num(v):
    """Coerce a cell to float; comma-decimal safe; None on failure."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _month_of(v):
    """Extract month (1-12) from a date-ish cell."""
    if isinstance(v, datetime):
        return v.month
    if v is None:
        return None
    s = str(v).strip()
    # try common formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:19], fmt).month
        except ValueError:
            continue
    return None


def parse_historical_workbook(content: bytes):
    """
    Returns dict:
      {
        "year": 2024,
        "monthly": {1: {"receipts_kg":.., "count":..}, ...},
        "totals": {"receipts_kg":.., "dispatches_kg":.., "disposal_kg":.., "opening_stock":..},
      }
    Raises ValueError with a clear message on structural problems.
    """
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)

    # ── Find the year tab: a sheet named like a 4-digit year ──
    year_sheet = None
    for name in wb.sheetnames:
        if re.fullmatch(r"\d{4}", name.strip()):
            year_sheet = name.strip()
            break
    if not year_sheet:
        raise ValueError("No se encontró una pestaña de año (ej. '2024').")
    year = int(year_sheet)

    # ── Parse receipts tab: monthly receipt totals ──
    ws = wb[year_sheet]
    monthly = {}   # month -> {"receipts_kg":float, "count":int}
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # skip header
    for row in rows:
        if not row or len(row) < 5:
            continue
        col0, col1, col4 = row[0], row[1], row[4]
        idv = str(col1).strip() if col1 is not None else ""
        # Only TRUE receipts (skip interleaved entrance-lot subtotal rows)
        if not _RECEIPT_ID.match(idv):
            continue
        kg = _num(col4)
        if kg is None:
            continue
        m = _month_of(col0)
        if not m:
            continue
        slot = monthly.setdefault(m, {"receipts_kg": 0.0, "count": 0})
        slot["receipts_kg"] += kg
        slot["count"] += 1

    # ── Parse Mass Balance tab: dispatch/disposal/opening totals ──
    mb_name = None
    for name in wb.sheetnames:
        if name.strip().lower().startswith("mass balance"):
            mb_name = name
            break

    totals = {"receipts_kg": sum(v["receipts_kg"] for v in monthly.values()),
              "dispatches_kg": 0.0, "disposal_kg": 0.0, "opening_stock": None}

    if mb_name:
        ws = wb[mb_name]
        all_rows = list(ws.iter_rows(values_only=True))
        # data starts after the header block (~row 7, 0-indexed 6)
        for row in all_rows[7:]:
            if not row:
                continue
            # opening stock: col0 contains STOCK
            c0 = str(row[0]).strip() if row[0] is not None else ""
            if "STOCK" in c0.upper() and totals["opening_stock"] is None:
                totals["opening_stock"] = _num(row[3])
            # disposals: col6 == DESECHO..., col8 = qty
            c6 = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
            if "DESECHO" in c6.upper():
                q = _num(row[8]) if len(row) > 8 else None
                if q:
                    totals["disposal_kg"] += q
            # dispatches: col11 = SA lot, col14 = qty
            c11 = str(row[11]).strip() if len(row) > 11 and row[11] is not None else ""
            if _DISPATCH_LOT.match(c11):
                q = _num(row[14]) if len(row) > 14 else None
                if q:
                    totals["dispatches_kg"] += q

    return {"year": year, "monthly": monthly, "totals": totals}


# ── Endpoint ────────────────────────────────────────────────────────────────
# @router.post("/historical/import")

async def import_historical_mass_balance(file, db, replace: bool = True):
    """
    Parse + store one year's historical summary.
    `replace=True` overwrites any existing rows for that year (idempotent re-import).
    """
    from models.historical_summary import HistoricalMonthlySummary, HistoricalYearMeta

    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("El archivo debe ser un Excel (.xlsx).")
    content = await file.read()
    parsed = parse_historical_workbook(content)

    year = parsed["year"]

    # Idempotent: clear existing rows for this year if replacing
    if replace:
        db.query(HistoricalMonthlySummary).filter(
            HistoricalMonthlySummary.year == year).delete()
        db.query(HistoricalYearMeta).filter(
            HistoricalYearMeta.year == year).delete()

    for m, vals in sorted(parsed["monthly"].items()):
        db.add(HistoricalMonthlySummary(
            year=year, month=m,
            receipts_kg=round(vals["receipts_kg"], 1),
            receipts_count=vals["count"],
            dispatches_kg=0, disposal_kg=0,   # monthly split not reliable; kept at year level
            source_file=file.filename,
        ))

    t = parsed["totals"]
    db.add(HistoricalYearMeta(
        year=year,
        opening_stock_kg=t["opening_stock"],
        total_receipts_kg=round(t["receipts_kg"], 1),
        total_dispatches_kg=round(t["dispatches_kg"], 1),
        total_disposal_kg=round(t["disposal_kg"], 1),
        source_file=file.filename,
    ))

    db.commit()

    return {
        "year": year,
        "months_imported": len(parsed["monthly"]),
        "total_receipts_kg": round(t["receipts_kg"], 1),
        "total_dispatches_kg": round(t["dispatches_kg"], 1),
        "total_disposal_kg": round(t["disposal_kg"], 1),
        "opening_stock_kg": t["opening_stock"],
        "monthly_receipts": {m: round(v["receipts_kg"], 1)
                             for m, v in sorted(parsed["monthly"].items())},
    }


@router.post("/historical/import")
async def historical_import_route(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
    replace: bool = True,
):
    try:
        return await import_historical_mass_balance(file, db, replace=replace)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    

@router.get("/historical/years")
def list_historical_years(db: Session = Depends(get_db)):
    """All imported years with their headline totals, newest first."""
    metas = db.query(HistoricalYearMeta).order_by(
        HistoricalYearMeta.year.desc()).all()
 
    return {
        "total": len(metas),
        "years": [
            {
                "year": m.year,
                "opening_stock_kg": m.opening_stock_kg,
                "total_receipts_kg": m.total_receipts_kg,
                "total_dispatches_kg": m.total_dispatches_kg,
                "total_disposal_kg": m.total_disposal_kg,
                "source_file": m.source_file,
                "imported_at": m.imported_at.isoformat() if m.imported_at else None,
            }
            for m in metas
        ],
    }
 
 
@router.get("/historical/{year}")
def get_historical_year(year: int, db: Session = Depends(get_db)):
    """One year: monthly receipt breakdown + yearly totals."""
    meta = db.query(HistoricalYearMeta).filter(
        HistoricalYearMeta.year == year).first()
    if not meta:
        raise HTTPException(status_code=404,
            detail=f"No hay datos históricos importados para {year}.")
 
    months = db.query(HistoricalMonthlySummary).filter(
        HistoricalMonthlySummary.year == year
    ).order_by(HistoricalMonthlySummary.month).all()
 
    return {
        "year": year,
        "opening_stock_kg": meta.opening_stock_kg,
        "total_receipts_kg": meta.total_receipts_kg,
        "total_dispatches_kg": meta.total_dispatches_kg,
        "total_disposal_kg": meta.total_disposal_kg,
        "source_file": meta.source_file,
        "months": [
            {
                "month": m.month,
                "receipts_kg": m.receipts_kg,
                "receipts_count": m.receipts_count,
            }
            for m in months
        ],
    }
 
 
@router.delete("/historical/{year}", status_code=204)
def delete_historical_year(
    year: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Remove an imported year (e.g. to re-import a corrected file)."""
    deleted = db.query(HistoricalMonthlySummary).filter(
        HistoricalMonthlySummary.year == year).delete()
    db.query(HistoricalYearMeta).filter(
        HistoricalYearMeta.year == year).delete()
    db.commit()
    if deleted == 0:
        raise HTTPException(status_code=404,
            detail=f"No hay datos históricos para {year}.")
    return None