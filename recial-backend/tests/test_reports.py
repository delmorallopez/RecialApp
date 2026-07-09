"""
test_reports.py — Step 4: report & file generation returns VALID output.

The core idea: a valid file has a recognisable signature in its first bytes.
  - XLSX is a ZIP archive → starts with b'PK'
  - PDF                    → starts with b'%PDF-'
So we hit each endpoint and assert the bytes are a real, well-formed file.
This is exactly what catches the bugs you hit:
  - reportlab not installed  → endpoint 500s (caught: status != 200)
  - the :.6f coordinate crash → truncated/invalid PDF (caught: no %PDF- / unopenable)
  - malformed workbook       → not a valid zip (caught: openpyxl can't reload it)

These report endpoints take a `year`/`supplier_id` and (per the routes shown)
require NO auth. If yours do require a token, add `headers=auth_headers`.

Endpoints needing real data (traceability, urban) build their chain first.
"""

import io
import zipfile
import pytest
from openpyxl import load_workbook

from auth import hash_password
from models.users import User, UserRole


TEST_USERNAME = "testadmin"
TEST_PASSWORD = "TestPass123!"


@pytest.fixture()
def test_user(db_session):
    user = User(
        username=TEST_USERNAME, full_name="Test Admin",
        email="testadmin@example.com",
        hashed_password=hash_password(TEST_PASSWORD),
        role=UserRole.ADMIN, is_active=True,
    )
    db_session.add(user)
    db_session.commit()
    db_session.refresh(user)
    return user


@pytest.fixture()
def auth_headers(client, test_user):
    res = client.post("/auth/login", data={
        "username": TEST_USERNAME, "password": TEST_PASSWORD,
    })
    assert res.status_code == 200, res.text
    return {"Authorization": f"Bearer {res.json()['access_token']}"}


# ── Reusable assertions ─────────────────────────────────────────────────────
def assert_valid_xlsx(res):
    """The response is a real, openable .xlsx workbook."""
    assert res.status_code == 200, res.text
    content = res.content
    assert content[:2] == b"PK", "Not a ZIP/XLSX (bad magic bytes)"
    # Prove it actually opens as a workbook, not just that it's a zip.
    wb = load_workbook(io.BytesIO(content))
    assert len(wb.sheetnames) >= 1


def assert_valid_pdf(res):
    """The response is a real PDF (starts with %PDF- and isn't truncated)."""
    assert res.status_code == 200, res.text
    content = res.content
    assert content[:5] == b"%PDF-", "Not a valid PDF (missing %PDF- header)"
    assert content.rstrip().endswith(b"%%EOF"), "PDF looks truncated (no %%EOF)"


# ════════════════════════════════════════════════════════════════════════════
# EXCEL REPORTS — work on empty data, no prerequisites needed
# ════════════════════════════════════════════════════════════════════════════

def test_mass_balance_excel_valid(client):
    """Mass-balance Excel returns a valid workbook (path confirmed: /reports/mass-balance)."""
    res = client.get("/reports/mass-balance?year=2026")
    assert_valid_xlsx(res)


def test_quarterly_excel_valid(client):
    """Quarterly closing Excel returns a valid workbook (path confirmed from route)."""
    res = client.get("/reports/quarterly-closing/excel?year=2026")
    assert_valid_xlsx(res)


def test_quarterly_excel_has_expected_sheet(client):
    """Sanity: the quarterly workbook has the CIERRES TRIMESTRALES sheet."""
    res = client.get("/reports/quarterly-closing/excel?year=2026")
    assert res.status_code == 200, res.text
    wb = load_workbook(io.BytesIO(res.content))
    assert "CIERRES TRIMESTRALES" in wb.sheetnames


# ════════════════════════════════════════════════════════════════════════════
# PDF REPORTS — need real data; see helpers below.
# Paths marked ⚠️ VERIFY — confirm against your traceability / reports routes.
# ════════════════════════════════════════════════════════════════════════════

# The following are scaffolded but SKIPPED until you confirm the exact paths
# and the create endpoints for building the chain. Remove the skip once wired.

# ── Chain builder: supplier → receipt → entrance → dispatch ─────────────────
def _build_full_chain(client, auth_headers):
    """
    Build a complete traceability chain and return the ids needed to trace it:
      supplier (Urban) → receipt → entrance (links receipt, type "B") → dispatch.
    Returns (receipt_id, dispatch_id).
    """
    # 1. Urban supplier (type "B" in entrance-speak)
    sup = client.post("/suppliers/", json={
        "supplier_type": "Urban",
        "name": "Chain Test Supplier",
        "cif": "B77777777",
    }, headers=auth_headers)
    assert sup.status_code in (200, 201), sup.text
    supplier_id = sup.json()["id"]

    # 2. receipt (unassigned)
    rec = client.post("/receipts/", json={
        "supplier_id": supplier_id,
        "date": "2026-01-10",
        "quantity_kg": 200.0,
        "raw_material": "UCO",
    }, headers=auth_headers)
    assert rec.status_code in (200, 201), rec.text
    receipt_id = rec.json()["id"]

    # 3. entrance linking that receipt. Urban → supplier_type "B".
    ent = client.post("/entrances/", json={
        "receipt_ids": [receipt_id],
        "supplier_type": "B",
        "date": "2026-01-11",
    }, headers=auth_headers)
    assert ent.status_code == 201, ent.text
    entrance_id = ent.json()["id"]

    # 4. customer + dispatch linking that entrance
    cust = client.post("/customers/", json={"name": "Chain Test Customer"},
                      headers=auth_headers)
    assert cust.status_code in (200, 201), cust.text
    customer_id = cust.json()["id"]

    disp = client.post("/dispatches/", json={
        "customer_id": customer_id,
        "date": "2026-01-15",
        "quantity": 150,
        "entrance_ids": [entrance_id],
    }, headers=auth_headers)
    assert disp.status_code == 201, disp.text
    dispatch_id = disp.json()["id"]

    return receipt_id, dispatch_id


def test_traceability_forward_pdf_valid(client, auth_headers):
    """Forward trace PDF (Receipt → Entrance → Dispatch) is a valid PDF."""
    receipt_id, _ = _build_full_chain(client, auth_headers)
    res = client.get(f"/traceability/forward/{receipt_id}/pdf")
    assert_valid_pdf(res)


def test_traceability_backward_pdf_valid(client, auth_headers):
    """Backward trace PDF (Dispatch → Entrance → Receipt) is a valid PDF."""
    _, dispatch_id = _build_full_chain(client, auth_headers)
    res = client.get(f"/traceability/backward/{dispatch_id}/pdf")
    assert_valid_pdf(res)


# ── Urban collection PDF — fully wired (builds Urban supplier + receipt) ─────
def test_urban_collection_pdf_valid(client, auth_headers):
    """
    Urban-collection PDF returns a valid PDF.
    This is the report whose :.6f coordinate crash produced an unopenable file,
    so the %PDF- + %%EOF checks in assert_valid_pdf are the real guard here.

    Builds the minimum data it needs: an Urban supplier with a pickup point and
    a receipt in the date range.
    """
    # 1. Urban supplier
    sup = client.post("/suppliers/", json={
        "supplier_type": "Urban",
        "name": "Urban PDF Test Supplier",
        "cif": "B99999999",
    }, headers=auth_headers)
    assert sup.status_code in (200, 201), sup.text
    supplier_id = sup.json()["id"]

    # 2. pickup point WITH coordinates (this is what triggered the old crash)
    pp = client.post("/pickup-points/", json={
        "supplier_id": supplier_id,
        "name": "Contenedor Centro",
        "latitude": 37.385100,
        "longitude": -4.173400,
    }, headers=auth_headers)
    assert pp.status_code in (200, 201), pp.text

    # 3. a receipt in range
    rec = client.post("/receipts/", json={
        "supplier_id": supplier_id,
        "date": "2026-01-10",
        "quantity_kg": 120.0,
        "raw_material": "UCO",
    }, headers=auth_headers)
    assert rec.status_code in (200, 201), rec.text

    # 4. generate the PDF (no auth required on this route)
    res = client.get(f"/reports/urban-collection/{supplier_id}/pdf"
                     "?date_from=2026-01-01&date_to=2026-12-31")
    assert_valid_pdf(res)