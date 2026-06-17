from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, extract
from datetime import date, datetime
from typing import Optional
from io import BytesIO

from database import get_db
from models.entrances import Entrance
from models.dispatches import Dispatch
from models.disposals import Disposal
from models.receipts  import Receipt
from models.suppliers import Supplier
from models.tanks     import Tank

router = APIRouter(prefix="/reports", tags=["Reports"])

MONTHS_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
             "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

QUARTER_MONTHS = {
    1: [("ENERO",1),("FEBRERO",2),("MARZO",3)],
    2: [("ABRIL",4),("MAYO",5),("JUNIO",6)],
    3: [("JULIO",7),("AGOSTO",8),("SEPTIEMBRE",9)],
    4: [("OCTUBRE",10),("NOVIEMBRE",11),("DICIEMBRE",12)],
}


# ════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════

def get_opening_stock(year: int, db: Session) -> float:
    prev_ent = db.query(func.sum(Entrance.quantity_kg)).filter(
        extract("year", Entrance.date) < year).scalar() or 0
    prev_dis = db.query(func.sum(Dispatch.quantity)).filter(
        extract("year", Dispatch.date) < year).scalar() or 0
    prev_mer = db.query(func.sum(Disposal.quantity)).join(
        Dispatch, Disposal.dispatch_id == Dispatch.id).filter(
        extract("year", Dispatch.date) < year).scalar() or 0
    return max(round(prev_ent - prev_dis - prev_mer, 1), 0)


def get_monthly_map(year: int, db: Session):
    """Returns (monthly_entrances, monthly_mermas, monthly_salidas) dicts keyed by month int."""
    ent_rows = db.query(
        extract("month", Entrance.date).label("m"),
        func.sum(Entrance.quantity_kg).label("kg")
    ).filter(extract("year", Entrance.date) == year).group_by("m").all()

    dis_rows = db.query(
        extract("month", Dispatch.date).label("m"),
        func.sum(Dispatch.quantity).label("kg")
    ).filter(extract("year", Dispatch.date) == year).group_by("m").all()

    mer_rows = db.query(
        extract("month", Dispatch.date).label("m"),
        func.sum(Disposal.quantity).label("kg")
    ).join(Dispatch, Disposal.dispatch_id == Dispatch.id).filter(
        extract("year", Dispatch.date) == year).group_by("m").all()

    monthly_ent  = {int(r.m): float(r.kg or 0) for r in ent_rows}
    monthly_dis  = {int(r.m): float(r.kg or 0) for r in dis_rows}
    monthly_mer  = {int(r.m): float(r.kg or 0) for r in mer_rows}
    return monthly_ent, monthly_mer, monthly_dis


# ════════════════════════════════════════════════════════════
# MASS BALANCE — Excel download
# ════════════════════════════════════════════════════════════

def excel_mass_balance(entrances, dispatches, year: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Mass Balance"

    GREEN = "2d7a4f"; AMBER = "d97706"; BLUE = "1d4ed8"
    DARK  = "1a1a2e"; GRAY  = "6b7280"; WHITE = "FFFFFF"

    def thin():
        s = Side(style="thin", color="D1D5DB")
        return Border(top=s, bottom=s, left=s, right=s)

    def hdr_cell(ws, row, col, value, bg, color=WHITE, bold=True, halign="center", wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, color=color, size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
        c.border = thin()
        return c

    def data_cell(ws, row, col, value, bg=WHITE, bold=False, halign="center", num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, color=DARK, size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=halign, vertical="center")
        c.border = thin()
        if num_fmt: c.number_format = num_fmt
        return c

    widths = {"A":16,"B":12,"C":12,"D":14,"E":10,"F":2,"G":18,"H":12,"I":12,"J":10,"K":2,"L":14,"M":12,"N":12,"O":14,"P":10,"Q":10,"R":12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value = f"RECICLAJES RECIAL S.L. — BALANCE DE MASAS {year}"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=13)
    c.fill = PatternFill("solid", start_color=GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("L2:R2")
    ws["L2"].value = "PG.09.01 / REG-A Balance de Masas"
    ws["L2"].font = Font(name="Arial", color=GRAY, size=9)
    ws["L2"].alignment = Alignment(horizontal="right")
    ws.row_dimensions[3].height = 6

    ws.row_dimensions[4].height = 22
    for rng, label, bg in [("A4:E4","ENTRADAS",GREEN),("G4:J4","MERMAS (PROCESO INTERNO)",AMBER),("L4:R4","SALIDAS",BLUE)]:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = label
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[5].height = 30
    for col, lbl in [(1,"LOTE"),(2,"FECHA"),(3,"CONCEPTO"),(4,"CANTIDAD Kg"),(5,"VALOR GEI")]:
        hdr_cell(ws, 5, col, lbl, GREEN, wrap=True)
    for col, lbl in [(7,"CONCEPTO"),(8,"FECHA"),(9,"CANTIDAD Kg"),(10,"VALOR GEI")]:
        hdr_cell(ws, 5, col, lbl, AMBER, wrap=True)
    for col, lbl in [(12,"LOTE"),(13,"CONCEPTO"),(14,"FECHA"),(15,"CANTIDAD Kg Netos"),(16,"VALOR GEI"),(17,"Nº POS"),(18,"STOCK")]:
        hdr_cell(ws, 5, col, lbl, BLUE, wrap=True)

    disposals = [d.disposal for d in dispatches if d.disposal]
    row = 6
    max_rows = max(len(entrances), len(disposals), len(dispatches))

    for i in range(max_rows):
        bg_e = "F8FAFC" if i % 2 == 0 else WHITE
        bg_m = "FFFBEB" if i % 2 == 0 else "FEF9EE"
        bg_s = "EFF6FF" if i % 2 == 0 else "F0F7FF"

        if i < len(entrances):
            e = entrances[i]
            data_cell(ws,row,1,e.batch_id,bg_e,bold=True)
            data_cell(ws,row,2,str(e.date),bg_e,num_fmt="DD/MM/YYYY")
            data_cell(ws,row,3,"UCO",bg_e,halign="left")
            data_cell(ws,row,4,int(e.quantity_kg or 0),bg_e,num_fmt="#,##0")
            data_cell(ws,row,5,e.value_gei or 1,bg_e)
        if i < len(disposals):
            m = disposals[i]
            data_cell(ws,row,7,"DESECHO SOLIDO",bg_m,halign="left")
            data_cell(ws,row,8,str(m.date),bg_m,num_fmt="DD/MM/YYYY")
            data_cell(ws,row,9,m.quantity,bg_m,num_fmt="#,##0")
            data_cell(ws,row,10,1,bg_m)
        if i < len(dispatches):
            d = dispatches[i]
            data_cell(ws,row,12,d.batch_id,bg_s,bold=True)
            data_cell(ws,row,13,d.raw_material or "UCO",bg_s)
            data_cell(ws,row,14,str(d.date),bg_s,num_fmt="DD/MM/YYYY")
            data_cell(ws,row,15,d.quantity,bg_s,num_fmt="#,##0")
            data_cell(ws,row,16,d.value_gei or 1,bg_s)
            data_cell(ws,row,17,d.post_number or "",bg_s)
            data_cell(ws,row,18,0,bg_s,num_fmt="#,##0")
        row += 1

    ws.row_dimensions[row].height = 22
    total_ent = sum(int(e.quantity_kg or 0) for e in entrances)
    total_mer = sum(m.quantity for m in disposals)
    total_sal = sum(d.quantity for d in dispatches)

    ws.merge_cells(f"A{row}:C{row}")
    ws.merge_cells(f"G{row}:H{row}")
    ws.merge_cells(f"L{row}:N{row}")

    for col, val, bg in [(1,"TOTAL",GREEN),(4,total_ent,GREEN),(7,"TOTAL",AMBER),(9,total_mer,AMBER),(12,"TOTAL",BLUE),(15,total_sal,BLUE)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
        if isinstance(val, int): c.number_format = "#,##0"

    ws.freeze_panes = "A6"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/mass-balance")
def get_mass_balance(
    year: int = Query(default=2024, ge=2020, le=2030),
    db: Session = Depends(get_db),
):
    entrances = db.query(Entrance).filter(
        Entrance.date >= date(year, 1, 1),
        Entrance.date <= date(year, 12, 31),
    ).order_by(Entrance.date).all()

    dispatches = db.query(Dispatch).options(
        joinedload(Dispatch.disposal)
    ).filter(
        Dispatch.date >= date(year, 1, 1),
        Dispatch.date <= date(year, 12, 31),
    ).order_by(Dispatch.date).all()

    excel_bytes = excel_mass_balance(entrances, dispatches, year)
    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=MassBalance_Recial_{year}.xlsx"}
    )


# ════════════════════════════════════════════════════════════
# RECEIPTS SUMMARY
# ════════════════════════════════════════════════════════════

@router.get("/receipts-summary")
def get_receipts_summary(
    date_from: Optional[date] = Query(None),
    date_to:   Optional[date] = Query(None),
    supplier_type: Optional[str] = Query(None),
    supplier_id:   Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    query = db.query(Receipt).join(
        Supplier, Receipt.supplier_id == Supplier.id
    ).options(joinedload(Receipt.supplier))

    if date_from:    query = query.filter(Receipt.date >= date_from)
    if date_to:      query = query.filter(Receipt.date <= date_to)
    if supplier_type: query = query.filter(Supplier.supplier_type == supplier_type)
    if supplier_id:  query = query.filter(Receipt.supplier_id == supplier_id)

    receipts = query.order_by(Receipt.date.desc()).all()

    supplier_summary = {}
    for r in receipts:
        sid = r.supplier_id
        if sid not in supplier_summary:
            supplier_summary[sid] = {
                "supplier_id":   sid,
                "supplier_name": r.supplier.name if r.supplier else "—",
                "supplier_type": r.supplier.supplier_type if r.supplier else "—",
                "receipts_count": 0, "total_kg": 0,
                "first_date": str(r.date), "last_date": str(r.date),
            }
        s = supplier_summary[sid]
        s["receipts_count"] += 1
        s["total_kg"] += r.quantity_kg or 0
        if str(r.date) < s["first_date"]: s["first_date"] = str(r.date)
        if str(r.date) > s["last_date"]:  s["last_date"]  = str(r.date)

    horeca_kg = sum(r.quantity_kg or 0 for r in receipts if r.supplier and r.supplier.supplier_type == "Horeca")
    urban_kg  = sum(r.quantity_kg or 0 for r in receipts if r.supplier and r.supplier.supplier_type == "Urban")

    return {
        "total_receipts": len(receipts),
        "total_kg":  round(sum(r.quantity_kg or 0 for r in receipts), 1),
        "horeca_kg": round(horeca_kg, 1),
        "urban_kg":  round(urban_kg, 1),
        "suppliers": sorted(
            [{**v,"total_kg":round(v["total_kg"],1)} for v in supplier_summary.values()],
            key=lambda x: x["total_kg"], reverse=True
        ),
    }


# ════════════════════════════════════════════════════════════
# TANK STOCK
# ════════════════════════════════════════════════════════════

@router.get("/tank-stock")
def get_tank_stock(
    year: int = Query(default=None),
    db: Session = Depends(get_db),
):
    if not year: year = datetime.now().year
    tanks = db.query(Tank).filter(Tank.is_active == True).all()
    result = []

    for tank in tanks:
        pct = round((tank.stock / tank.capacity) * 100, 1) if tank.capacity else 0

        ent_rows = db.query(
            extract("month", Entrance.date).label("m"),
            func.sum(Entrance.quantity_kg).label("t")
        ).filter(
            Entrance.tank_id == tank.id,
            extract("year", Entrance.date) == year
        ).group_by(extract("month", Entrance.date)).all()

        dis_rows = db.query(
            extract("month", Dispatch.date).label("m"),
            func.sum(Dispatch.quantity).label("t")
        ).filter(
            Dispatch.tank_id == tank.id,
            extract("year", Dispatch.date) == year
        ).group_by(extract("month", Dispatch.date)).all()

        mer_rows = db.query(
            extract("month", Dispatch.date).label("m"),
            func.sum(Disposal.quantity).label("t")
        ).join(
            Dispatch, Disposal.dispatch_id == Dispatch.id
        ).filter(
            Dispatch.tank_id == tank.id,
            extract("year", Dispatch.date) == year
        ).group_by(extract("month", Dispatch.date)).all()

        monthly_in  = {int(r[0]): float(r[1] or 0) for r in ent_rows}
        monthly_out = {int(r[0]): float(r[1] or 0) for r in dis_rows}
        monthly_mer = {int(r[0]): float(r[1] or 0) for r in mer_rows}

        monthly = []
        running = 0
        for m in range(1, 13):
            added   = monthly_in.get(m, 0)
            removed = monthly_out.get(m, 0) + monthly_mer.get(m, 0)
            running = running + added - removed
            fp = round((running/tank.capacity)*100,1) if tank.capacity else None
            monthly.append({"month":m,"label":["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][m-1],"added":round(added,1),"removed":round(removed,1),"stock":round(max(running,0),1),"fill_pct":fp})

        result.append({
            "id":tank.id,"name":tank.name,"capacity":tank.capacity or 0,
            "current_stock":tank.stock or 0,"current_pct":pct,
            "monthly":monthly,
            "total_in":round(sum(monthly_in.values()),1),
            "total_out":round(sum(monthly_out.values())+sum(monthly_mer.values()),1),
        })

    return {"year":year,"tanks":result}


# ════════════════════════════════════════════════════════════
# DISPATCHES SUMMARY
# ════════════════════════════════════════════════════════════

@router.get("/dispatches-summary")
def get_dispatches_summary(
    date_from:   Optional[date] = Query(None),
    date_to:     Optional[date] = Query(None),
    customer_id: Optional[int]  = Query(None),
    db: Session = Depends(get_db),
):
    from models.customers import Customer
    query = db.query(Dispatch).options(
        joinedload(Dispatch.customer), joinedload(Dispatch.disposal))
    if date_from:   query = query.filter(Dispatch.date >= date_from)
    if date_to:     query = query.filter(Dispatch.date <= date_to)
    if customer_id: query = query.filter(Dispatch.customer_id == customer_id)
    dispatches = query.order_by(Dispatch.date.desc()).all()

    customer_summary = {}
    for d in dispatches:
        cid = d.customer_id
        if cid not in customer_summary:
            customer_summary[cid] = {
                "customer_id":d.customer_id,
                "customer_name":d.customer.name if d.customer else "—",
                "dispatches_count":0,"total_kg":0,"total_disposal_kg":0,
                "first_date":str(d.date),"last_date":str(d.date),
            }
        s = customer_summary[cid]
        s["dispatches_count"]  += 1
        s["total_kg"]          += d.quantity or 0
        s["total_disposal_kg"] += d.disposal.quantity if d.disposal else 0
        if str(d.date) < s["first_date"]: s["first_date"] = str(d.date)
        if str(d.date) > s["last_date"]:  s["last_date"]  = str(d.date)

    monthly_map = {}
    for d in dispatches:
        key = str(d.date)[:7]
        if key not in monthly_map:
            monthly_map[key] = {"month":key,"label":d.date.strftime("%b %Y") if hasattr(d.date,"strftime") else key,"dispatched":0,"disposal":0,"count":0}
        monthly_map[key]["dispatched"] += d.quantity or 0
        monthly_map[key]["disposal"]   += d.disposal.quantity if d.disposal else 0
        monthly_map[key]["count"]      += 1

    total_kg          = sum(d.quantity or 0 for d in dispatches)
    total_disposal_kg = sum(d.disposal.quantity if d.disposal else 0 for d in dispatches)

    return {
        "total_dispatches":len(dispatches),"total_kg":round(total_kg,1),"total_disposal_kg":round(total_disposal_kg,1),
        "monthly":sorted([{**v,"dispatched":round(v["dispatched"],1),"disposal":round(v["disposal"],1)} for v in monthly_map.values()],key=lambda x:x["month"]),
        "customers":sorted([{**v,"total_kg":round(v["total_kg"],1),"total_disposal_kg":round(v["total_disposal_kg"],1)} for v in customer_summary.values()],key=lambda x:x["total_kg"],reverse=True),
    }


# ════════════════════════════════════════════════════════════
# CUSTOMER ACTIVITY
# ════════════════════════════════════════════════════════════

@router.get("/customer-activity")
def get_customer_activity(
    date_from:    Optional[date]  = Query(None),
    date_to:      Optional[date]  = Query(None),
    customer_id:  Optional[int]   = Query(None),
    price_per_kg: float           = Query(default=1.09),
    db: Session = Depends(get_db),
):
    query = db.query(Dispatch).options(joinedload(Dispatch.customer),joinedload(Dispatch.disposal))
    if date_from:   query = query.filter(Dispatch.date >= date_from)
    if date_to:     query = query.filter(Dispatch.date <= date_to)
    if customer_id: query = query.filter(Dispatch.customer_id == customer_id)
    dispatches = query.order_by(Dispatch.date).all()
    today = date.today()

    customer_map = {}
    for d in dispatches:
        cid = d.customer_id
        if cid not in customer_map:
            customer_map[cid] = {"customer_id":cid,"customer_name":d.customer.name if d.customer else "—","dispatches":[],"monthly":{}}
        c = customer_map[cid]
        c["dispatches"].append(d)
        key = str(d.date)[:7]
        if key not in c["monthly"]:
            c["monthly"][key] = {"month":key,"label":d.date.strftime("%b %Y") if hasattr(d.date,"strftime") else key,"kg":0}
        c["monthly"][key]["kg"] += d.quantity or 0

    result = []
    for cid, data in customer_map.items():
        disps     = data["dispatches"]
        total_kg  = sum(d.quantity or 0 for d in disps)
        revenue   = round(total_kg * price_per_kg, 2)
        avg_kg    = round(total_kg / len(disps), 1) if disps else 0
        last_date = max(str(d.date) for d in disps)
        first_date= min(str(d.date) for d in disps)
        days_since= (today - date.fromisoformat(last_date)).days
        status    = "active" if days_since<=60 else "inactive" if days_since<=180 else "dormant"
        monthly_sorted = sorted(data["monthly"].values(), key=lambda x:x["kg"],reverse=True)
        best_month = monthly_sorted[0]["label"] if monthly_sorted else "—"
        result.append({
            "customer_id":cid,"customer_name":data["customer_name"],
            "dispatches_count":len(disps),"total_kg":round(total_kg,1),
            "revenue":revenue,"avg_kg_per_dispatch":avg_kg,
            "first_date":first_date,"last_date":last_date,
            "days_since_last":days_since,"status":status,"best_month":best_month,
            "monthly_trend":sorted(data["monthly"].values(),key=lambda x:x["month"]),
        })
    result.sort(key=lambda x:x["total_kg"],reverse=True)
    total_kg = sum(r["total_kg"] for r in result)
    return {"total_customers":len(result),"total_kg":round(total_kg,1),"total_revenue":round(total_kg*price_per_kg,2),"price_per_kg":price_per_kg,"customers":result}


# ════════════════════════════════════════════════════════════
# SUPPLIER ACTIVITY
# ════════════════════════════════════════════════════════════

@router.get("/supplier-activity")
def get_supplier_activity(
    date_from:     Optional[date] = Query(None),
    date_to:       Optional[date] = Query(None),
    supplier_id:   Optional[int]  = Query(None),
    supplier_type: Optional[str]  = Query(None),
    db: Session = Depends(get_db),
):
    query = db.query(Receipt).options(joinedload(Receipt.supplier))
    if date_from:     query = query.filter(Receipt.date >= date_from)
    if date_to:       query = query.filter(Receipt.date <= date_to)
    if supplier_id:   query = query.filter(Receipt.supplier_id == supplier_id)
    if supplier_type: query = query.join(Supplier,Receipt.supplier_id==Supplier.id).filter(Supplier.supplier_type==supplier_type)
    receipts = query.order_by(Receipt.date).all()
    today = date.today()

    supplier_map = {}
    for r in receipts:
        sid = r.supplier_id
        if sid not in supplier_map:
            supplier_map[sid] = {"supplier_id":sid,"supplier_name":r.supplier.name if r.supplier else "—","supplier_type":r.supplier.supplier_type if r.supplier else "—","receipts":[],"monthly":{}}
        s = supplier_map[sid]
        s["receipts"].append(r)
        key = str(r.date)[:7]
        if key not in s["monthly"]:
            s["monthly"][key] = {"month":key,"label":r.date.strftime("%b %Y") if hasattr(r.date,"strftime") else key,"kg":0,"count":0}
        s["monthly"][key]["kg"]    += r.quantity_kg or 0
        s["monthly"][key]["count"] += 1

    result = []
    for sid, data in supplier_map.items():
        recs       = data["receipts"]
        total_kg   = sum(r.quantity_kg or 0 for r in recs)
        avg_kg     = round(total_kg/len(recs),1) if recs else 0
        last_date  = max(str(r.date) for r in recs)
        first_date = min(str(r.date) for r in recs)
        days_since = (today - date.fromisoformat(last_date)).days
        status     = "active" if days_since<=60 else "inactive" if days_since<=180 else "dormant"
        monthly_sorted = sorted(data["monthly"].values(),key=lambda x:x["kg"],reverse=True)
        best_month = monthly_sorted[0]["label"] if monthly_sorted else "—"
        result.append({
            "supplier_id":sid,"supplier_name":data["supplier_name"],"supplier_type":data["supplier_type"],
            "receipts_count":len(recs),"total_kg":round(total_kg,1),"avg_kg_per_receipt":avg_kg,
            "first_date":first_date,"last_date":last_date,"days_since_last":days_since,
            "status":status,"best_month":best_month,
            "monthly_trend":sorted(data["monthly"].values(),key=lambda x:x["month"]),
        })
    result.sort(key=lambda x:x["total_kg"],reverse=True)
    total_kg  = sum(r["total_kg"] for r in result)
    horeca_kg = sum(r["total_kg"] for r in result if r["supplier_type"]=="Horeca")
    urban_kg  = sum(r["total_kg"] for r in result if r["supplier_type"]=="Urban")
    return {"total_suppliers":len(result),"total_kg":round(total_kg,1),"horeca_kg":round(horeca_kg,1),"urban_kg":round(urban_kg,1),"suppliers":result}


# ════════════════════════════════════════════════════════════
# QUARTERLY CLOSING — JSON + Excel
# ════════════════════════════════════════════════════════════

def _build_quarterly(year: int, db: Session) -> dict:
    monthly_ent, monthly_mer, monthly_dis = get_monthly_map(year, db)
    opening_stock = get_opening_stock(year, db)
    running = opening_stock
    quarters = {}

    for q in range(1, 5):
        q_open  = running
        months_data = []
        q_total = 0
        q_sal   = 0

        for month_name, month_num in QUARTER_MONTHS[q]:
            kg_brutos = round(monthly_ent.get(month_num, 0), 1)
            mermas    = round(monthly_mer.get(month_num, 0), 1)
            net       = round(kg_brutos - mermas, 1)
            q_total  += net
            q_sal    += monthly_dis.get(month_num, 0)
            months_data.append({"name":month_name,"month":month_num,"kg_brutos":kg_brutos,"mermas":mermas,"total":net})

        ending = max(round(q_open + q_total - q_sal, 1), 0)
        quarters[q] = {
            "quarter":q,"label":f"TRIMESTRE {q}º",
            "opening_stock":round(q_open,1),
            "months":months_data,
            "total_kg":round(q_total,1),
            "salidas":round(q_sal,1),
            "ending_stock":ending,
        }
        running = ending

    return {
        "year":year,
        "opening_stock":opening_stock,
        "quarters":quarters,
        "year_ending_stock":running,
        "total_entrances_kg":round(sum(monthly_ent.values()),1),
        "total_mermas_kg":round(sum(monthly_mer.values()),1),
        "total_salidas_kg":round(sum(monthly_dis.values()),1),
    }


def _quarterly_excel(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "CIERRES TRIMESTRALES"

    GREEN  = "2d7a4f"; DGREEN = "1e3d2a"; LGRAY = "F2F2F2"
    AMBER  = "FFF2CC"; WHITE  = "FFFFFF"; DARK  = "1a1a2e"
    RED    = "C00000"

    def bdr():
        s = Side(style="thin", color="D0D0D0")
        return Border(top=s, bottom=s, left=s, right=s)

    def c(row, col, value, bold=False, bg=None, color=DARK,
          align="center", wrap=False, num_fmt=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font      = Font(name="Arial", bold=bold, color=color, size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if bg: cell.fill = PatternFill("solid", start_color=bg)
        cell.border = bdr()
        if num_fmt: cell.number_format = num_fmt
        return cell

    year = data["year"]

    # Title
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value     = f"RECICLAJES RECIAL S.L. — CIERRES TRIMESTRALES {year}"
    t.font      = Font(name="Arial", bold=True, color=WHITE, size=13)
    t.fill      = PatternFill("solid", start_color=DGREEN)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 6

    # Q1 top-left, Q2 top-right, Q3 bottom-left, Q4 bottom-right
    BLOCKS = {1:(1,3), 2:(6,3), 3:(1,14), 4:(6,14)}

    for q_num, (sc, sr) in BLOCKS.items():
        q = data["quarters"][q_num]

        # Quarter header
        ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc+3)
        h = ws.cell(row=sr, column=sc, value=q["label"])
        h.font      = Font(name="Arial", bold=True, color=WHITE, size=11)
        h.fill      = PatternFill("solid", start_color=GREEN)
        h.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[sr].height = 20

        # Column headers
        r = sr + 1
        ws.row_dimensions[r].height = 28
        for ci, hdr in enumerate(["", "CANTIDAD Kg/Brutos", "Mermas", "Total"]):
            c(r, sc+ci, hdr, bold=True, bg=LGRAY, wrap=True)

        # Opening stock
        r += 1
        c(r, sc,   "STOCK", bold=True, align="left", bg=AMBER)
        c(r, sc+1, "",      bg=AMBER)
        c(r, sc+2, "",      bg=AMBER)
        c(r, sc+3, q["opening_stock"], bold=True, bg=AMBER, num_fmt="#,##0.0")

        # Months
        for month in q["months"]:
            r += 1
            ws.row_dimensions[r].height = 18
            c(r, sc,   month["name"],      align="left")
            c(r, sc+1, month["kg_brutos"], num_fmt="#,##0.0")
            c(r, sc+2, month["mermas"],    num_fmt="#,##0.0")
            c(r, sc+3, month["total"],     num_fmt="#,##0.0")

        # TOTAL row
        r += 1
        ws.row_dimensions[r].height = 18
        c(r, sc,   "", bg=LGRAY); c(r, sc+1, "", bg=LGRAY)
        c(r, sc+2, "TOTAL",     bold=True, bg=LGRAY)
        c(r, sc+3, q["total_kg"], bold=True, bg=LGRAY, num_fmt="#,##0.0")

        # SALIDAS row
        r += 1
        ws.row_dimensions[r].height = 18
        c(r, sc,   "", bg=LGRAY); c(r, sc+1, "", bg=LGRAY)
        c(r, sc+2, "SALIDAS",    bold=True, bg=LGRAY, color=RED)
        c(r, sc+3, q["salidas"], bold=True, bg=LGRAY, color=RED, num_fmt="#,##0.0")

        # Ending stock
        r += 1
        ws.row_dimensions[r].height = 18
        c(r, sc,   "", bg=LGRAY); c(r, sc+1, "", bg=LGRAY)
        c(r, sc+2, "STOCK FINAL", bold=True, bg=AMBER)
        c(r, sc+3, q["ending_stock"], bold=True, bg=AMBER, num_fmt="#,##0.0")

    # Year summary row
    ws.row_dimensions[25].height = 8
    ws.merge_cells("A26:I26")
    s = ws["A26"]
    s.value = (
        f"STOCK 1 ENE {year}: {data['opening_stock']:,.0f} kg    "
        f"ENTRADAS: {data['total_entrances_kg']:,.0f} kg    "
        f"MERMAS: {data['total_mermas_kg']:,.0f} kg    "
        f"SALIDAS: {data['total_salidas_kg']:,.0f} kg    "
        f"STOCK FINAL {year}: {data['year_ending_stock']:,.0f} kg"
    )
    s.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
    s.fill      = PatternFill("solid", start_color=DGREEN)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[26].height = 22

    # Column widths
    for col, w in {1:16, 2:20, 3:10, 4:13, 5:3, 6:16, 7:20, 8:10, 9:13}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/quarterly-closing")
def get_quarterly_closing(
    year: int = Query(default=None),
    db: Session = Depends(get_db),
):
    if not year: year = datetime.now().year
    return _build_quarterly(year, db)


@router.get("/quarterly-closing/excel")
def download_quarterly_closing(
    year: int = Query(default=None),
    db: Session = Depends(get_db),
):
    if not year: year = datetime.now().year
    data  = _build_quarterly(year, db)
    excel = _quarterly_excel(data)
    return StreamingResponse(
        BytesIO(excel),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=CierresTrimestrales_{year}.xlsx"}
    )


# ════════════════════════════════════════════════════════════
# ANNUAL SUMMARY
# ════════════════════════════════════════════════════════════

@router.get("/annual-summary")
def get_annual_summary(
    year: int = Query(default=None),
    db: Session = Depends(get_db),
):
    if not year: year = datetime.now().year

    monthly_ent, monthly_mer, monthly_dis = get_monthly_map(year, db)
    opening_stock = get_opening_stock(year, db)
    running = opening_stock
    monthly = []

    for m in range(1, 13):
        ent_kg = round(monthly_ent.get(m, 0), 1)
        mer_kg = round(monthly_mer.get(m, 0), 1)
        dis_kg = round(monthly_dis.get(m, 0), 1)
        net    = round(ent_kg - mer_kg, 1)
        running = max(round(running + net - dis_kg, 1), 0)
        monthly.append({
            "month":m,"label":MONTHS_ES[m-1],
            "entrances_kg":ent_kg,"mermas_kg":mer_kg,
            "net_kg":net,"dispatches_kg":dis_kg,"stock":running,
        })

    horeca_kg = db.query(func.sum(Receipt.quantity_kg)).join(
        Supplier, Receipt.supplier_id==Supplier.id
    ).filter(Supplier.supplier_type=="Horeca",extract("year",Receipt.date)==year).scalar() or 0

    urban_kg = db.query(func.sum(Receipt.quantity_kg)).join(
        Supplier, Receipt.supplier_id==Supplier.id
    ).filter(Supplier.supplier_type=="Urban",extract("year",Receipt.date)==year).scalar() or 0

    return {
        "year":year,
        "opening_stock":opening_stock,
        "ending_stock":running,
        "total_entrances_kg":round(sum(monthly_ent.values()),1),
        "total_mermas_kg":round(sum(monthly_mer.values()),1),
        "total_dispatches_kg":round(sum(monthly_dis.values()),1),
        "horeca_kg":round(horeca_kg,1),
        "urban_kg":round(urban_kg,1),
        "monthly":monthly,
    }


# ════════════════════════════════════════════════════════════
# URBAN COLLECTION
# ════════════════════════════════════════════════════════════

def _build_urban_collection_data(supplier_id: int, date_from, date_to, db) -> dict:
    from models.receipt_pickup import ReceiptPickup
    from models.pickupPoints import PickupPoint
    from fastapi import HTTPException

    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    pickup_points = db.query(PickupPoint).filter(
        PickupPoint.supplier_id == supplier_id
    ).order_by(PickupPoint.id).all()

    query = db.query(Receipt).filter(Receipt.supplier_id == supplier_id)
    if date_from: query = query.filter(Receipt.date >= date_from)
    if date_to:   query = query.filter(Receipt.date <= date_to)
    receipts = query.order_by(Receipt.date).all()

    rows = []
    for r in receipts:
        from models.receipt_pickup import ReceiptPickup
        receipt_pickups = db.query(ReceiptPickup).filter(
            ReceiptPickup.receipt_id == r.id
        ).all()
        qty_map = {rp.pickup_point_id: rp.quantity_kg for rp in receipt_pickups}
        quantities = []
        for pp in pickup_points:
            qty = qty_map.get(pp.id)
            quantities.append(round(qty, 1) if qty else None)
        date_str = r.date.strftime("%d/%m/%y") if hasattr(r.date, "strftime") else str(r.date)
        rows.append({
            "date":       date_str,
            "receipt_id": r.id,
            "quantities": quantities,
            "total":      round(r.quantity_kg or 0, 1),
        })

    pp_totals = []
    for i in range(len(pickup_points)):
        col_total = sum(
            row["quantities"][i] for row in rows
            if row["quantities"][i] is not None
        )
        pp_totals.append(round(col_total, 1))

    grand_total = round(sum(r["total"] for r in rows), 1)
    totals = pp_totals + [grand_total]

    if date_from and date_to:
        period_label = (
            f"1º SEMESTRE {date_from.year}" if date_from.month <= 6
            else f"2º SEMESTRE {date_from.year}"
        )
    else:
        period_label = "PERÍODO PERSONALIZADO"

    return {
        "supplier_id":    supplier_id,
        "supplier_name":  supplier.name,
        "pickup_points":  [pp.name for pp in pickup_points],
        "rows":           rows,
        "totals":         totals,
        "period_label":   period_label,
        "date_from":      str(date_from) if date_from else None,
        "date_to":        str(date_to)   if date_to   else None,
        "total_receipts": len(rows),
        "grand_total_kg": grand_total,
    }


def _generate_urban_pdf(data: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    buf = BytesIO()
    w, h = A4
    c = canvas.Canvas(buf, pagesize=A4)

    LEFT  = 20 * mm
    RIGHT = w - 20 * mm
    TOP   = h - 18 * mm

    GREEN_DARK  = colors.HexColor("#3a7a54")
    GREEN_LIGHT = colors.HexColor("#8dc63f")
    GRAY_HDR    = colors.HexColor("#f2f2f2")
    WHITE       = colors.white

    supplier_name = data["supplier_name"]
    period_label  = data["period_label"]
    pickup_points = data["pickup_points"]
    rows          = data["rows"]
    totals        = data["totals"]
    n_pp          = len(pickup_points)

    y = TOP
    c.setFont("Helvetica-Bold", 15)
    c.setFillColor(GREEN_DARK)
    c.drawString(LEFT, y, "RESUMEN RECOGIDA URBANO")
    y -= 7 * mm
    c.setFont("Helvetica-Bold", 13)
    c.drawString(LEFT, y, f"LOCALIDAD: {supplier_name.upper()}")
    y -= 6 * mm
    c.drawString(LEFT, y, period_label.upper())

# ── Header right — Recial logo ────────────────────────────
    import os as _os
    _ASSETS   = _os.path.join(_os.path.dirname(__file__), "../assets")
    _LOGO     = _os.path.join(_ASSETS, "LogoRecial.png")
    logo_w = 45 * mm
    logo_h = 18 * mm
    logo_x = RIGHT - logo_w
    logo_y = TOP - logo_h + 6 * mm
    if _os.path.exists(_LOGO):
        c.drawImage(_LOGO, logo_x, logo_y,
                    width=logo_w, height=logo_h,
                    preserveAspectRatio=True, mask='auto')
    else:
        # Fallback green box if logo file missing
        c.setFillColor(GREEN_DARK)
        c.roundRect(logo_x, logo_y, logo_w, logo_h, 3*mm, fill=1, stroke=0)
        c.setFont("Helvetica-Bold", 18)
        c.setFillColor(WHITE)
        c.drawCentredString(logo_x + logo_w/2, logo_y + 5*mm, "recial")
        c.setFont("Helvetica", 7)
        c.drawCentredString(logo_x + logo_w/2, logo_y + 2*mm, "gestión de residuos")

    usable_w = RIGHT - LEFT
    fecha_w  = 26 * mm
    total_w  = 22 * mm
    pp_w     = (usable_w - fecha_w - total_w) / max(n_pp, 1)
    col_widths = [fecha_w] + [pp_w] * n_pp + [total_w]

    header1    = ["FECHA"] + ["CONTENEDORES"] + [""] * (n_pp - 1) + ["TOTAL"]
    header2    = [""] + [pp.upper() for pp in pickup_points] + [""]
    table_data = [header1, header2]

    for row in rows:
        r = [row["date"]]
        for qty in row["quantities"]:
            r.append(str(int(qty)) if qty is not None and qty > 0 else ("0" if qty == 0 else ""))
        r.append(str(int(row["total"])) if row["total"] else "0")
        table_data.append(r)

    tot_row = [""]
    for t in totals[:-1]:
        tot_row.append(str(int(t)) if t else "")
    tot_row.append(str(int(totals[-1])) if totals[-1] else "0")
    table_data.append(tot_row)

    n_data  = len(rows)
    tot_idx = 2 + n_data

    tbl = Table(table_data, colWidths=col_widths)
    GREEN_DARK  = colors.HexColor("#1e3d2a")   # dark green — main headers
    GREEN_MID   = colors.HexColor("#2d7a4f")   # mid green — pickup point names
    GREEN_LIGHT = colors.HexColor("#8dc63f")   # lime green — totals row
    GRAY_ROW    = colors.HexColor("#f2f2f2")   # alternating row bg
    WHITE       = colors.white
    DARK        = colors.HexColor("#1a1a1a")

    tbl.setStyle(TableStyle([
        # ── Header row 1: FECHA | CONTENEDORES | TOTAL ────────
        ("BACKGROUND",    (0,0),  (0,0),     GREEN_DARK),   # FECHA
        ("BACKGROUND",    (1,0),  (n_pp,0),  GREEN_DARK),   # CONTENEDORES
        ("BACKGROUND",    (-1,0), (-1,0),    GREEN_DARK),   # TOTAL
        ("TEXTCOLOR",     (0,0),  (-1,0),    WHITE),
        ("FONTNAME",      (0,0),  (-1,0),    "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),  (-1,0),    10),
        ("ALIGN",         (0,0),  (-1,0),    "CENTER"),
        ("VALIGN",        (0,0),  (-1,0),    "MIDDLE"),
        ("SPAN",          (1,0),  (n_pp,0)),

        # ── Header row 2: pickup point names ──────────────────
        ("BACKGROUND",    (0,1),  (0,1),     GREEN_DARK),   # empty FECHA cell
        ("BACKGROUND",    (1,1),  (n_pp,1),  GREEN_DARK),    # pickup names — mid green
        ("BACKGROUND",    (-1,1), (-1,1),    GREEN_DARK),   # empty TOTAL cell
        ("TEXTCOLOR",     (0,1),  (0,1),     WHITE),
        ("TEXTCOLOR",     (1,1),  (n_pp,1),  WHITE),
        ("TEXTCOLOR",     (-1,1), (-1,1),    WHITE),
        ("FONTNAME",      (0,1),  (-1,1),    "Helvetica-Bold"),
        ("FONTSIZE",      (0,1),  (-1,1),    9),
        ("ALIGN",         (0,1),  (-1,1),    "CENTER"),
        ("VALIGN",        (0,1),  (-1,1),    "MIDDLE"),

        # ── Data rows — alternating white / light gray ─────────
        ("FONTNAME",      (0,2),  (-1,tot_idx-1), "Helvetica-Bold"),
        ("FONTSIZE",      (0,2),  (-1,tot_idx-1), 10),
        ("ALIGN",         (0,2),  (-1,tot_idx-1), "CENTER"),
        ("VALIGN",        (0,2),  (-1,tot_idx-1), "MIDDLE"),
        ("TEXTCOLOR",     (0,2),  (-1,tot_idx-1), DARK),
        ("ROWBACKGROUNDS",(0,2),  (-1,tot_idx-1), [WHITE, GRAY_ROW]),

        # ── Totals row — lime green ────────────────────────────
        ("BACKGROUND",    (0,tot_idx), (-1,tot_idx), GREEN_LIGHT),
        ("TEXTCOLOR",     (0,tot_idx), (-1,tot_idx), WHITE),
        ("FONTNAME",      (0,tot_idx), (-1,tot_idx), "Helvetica-Bold"),
        ("FONTSIZE",      (0,tot_idx), (-1,tot_idx), 11),
        ("ALIGN",         (0,tot_idx), (-1,tot_idx), "CENTER"),
        ("VALIGN",        (0,tot_idx), (-1,tot_idx), "MIDDLE"),

        # ── Grid ───────────────────────────────────────────────
        ("GRID",          (0,0),  (-1,-1),   0.5, colors.HexColor("#cccccc")),
        ("ROWHEIGHT",     (0,0),  (-1,-1),   9 * mm),

        # ── Left-align FECHA column ────────────────────────────
        ("ALIGN",         (0,2),  (0,tot_idx), "CENTER"),
    ]))

    tbl_w, tbl_h = tbl.wrapOn(c, RIGHT - LEFT, h)
    tbl.drawOn(c, LEFT, y - 14 * mm - tbl_h)
    c.save()
    buf.seek(0)
    return buf.read()


@router.get("/urban-collection/{supplier_id}")
def get_urban_collection(
    supplier_id: int,
    date_from: Optional[date] = Query(None),
    date_to:   Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    return _build_urban_collection_data(supplier_id, date_from, date_to, db)


@router.get("/urban-collection/{supplier_id}/pdf")
def download_urban_collection_pdf(
    supplier_id:  int,
    date_from:    Optional[date] = Query(None),
    date_to:      Optional[date] = Query(None),
    period_label: Optional[str]  = Query(None),
    db: Session = Depends(get_db),
):
    data = _build_urban_collection_data(supplier_id, date_from, date_to, db)
    if period_label:
        data["period_label"] = period_label
    pdf_bytes = _generate_urban_pdf(data)
    filename  = f"RecogidaUrbano_{data['supplier_name'].replace(' ','_')}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/suppliers-list/pdf")
def get_suppliers_list_pdf(
    supplier_type: Optional[str] = Query(None),  # "Horeca" or "Urban"
    supplier_id:   Optional[int] = Query(None),  # specific supplier
    db: Session = Depends(get_db),
):
    from models.suppliers import Supplier
    from sqlalchemy.orm import joinedload

    query = db.query(Supplier).options(joinedload(Supplier.pickup_points))
    if supplier_type: query = query.filter(Supplier.supplier_type == supplier_type)
    if supplier_id:   query = query.filter(Supplier.id == supplier_id)
    suppliers = query.order_by(Supplier.supplier_type, Supplier.name).all()

    data = []
    for s in suppliers:
        pps = [{"name": pp.name, "latitude": pp.latitude, "longitude": pp.longitude, "notes": getattr(pp, "notes", "")}
               for pp in (s.pickup_points or [])]
        data.append({
                "name":          s.name,
                "supplier_type": s.supplier_type,
                "cif":           getattr(s, "cif", None),
                "address":       getattr(s, "address", None),
                "email":         getattr(s, "email", None),
                "phone":         getattr(s, "phone", None),
                "pickup_points": pps,
            })

    pdf = _generate_suppliers_pdf(data)
    return StreamingResponse(BytesIO(pdf), media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=Recial_Suppliers.pdf"})


@router.get("/suppliers-list")
def get_suppliers_list(
    supplier_type: Optional[str] = Query(None),
    supplier_id:   Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    from models.suppliers import Supplier
    from sqlalchemy.orm import joinedload

    query = db.query(Supplier).options(joinedload(Supplier.pickup_points))
    if supplier_type: query = query.filter(Supplier.supplier_type == supplier_type)
    if supplier_id:   query = query.filter(Supplier.id == supplier_id)
    suppliers = query.order_by(Supplier.supplier_type, Supplier.name).all()

    result = []
    for s in suppliers:
        pps = [{"id": pp.id, "name": pp.name, "latitude": pp.latitude, "longitude": pp.longitude}
               for pp in (s.pickup_points or [])]
        result.append({
                "id":            s.id,
                "name":          s.name,
                "supplier_type": s.supplier_type,
                "cif":           getattr(s, "cif", None),
                "address":       getattr(s, "address", None),
                "email":         getattr(s, "email", None),
                "phone":         getattr(s, "phone", None),
                "pickup_points": pps,
            })

    horeca = [r for r in result if r["supplier_type"] == "Horeca"]
    urban  = [r for r in result if r["supplier_type"] == "Urban"]
    return {"total": len(result), "horeca_count": len(horeca),
            "urban_count": len(urban), "suppliers": result}



# ════════════════════════════════════════════════════════════
# CUSTOMERS LIST
# ════════════════════════════════════════════════════════════

def _generate_customers_pdf(customers: list) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from datetime import date as date_type
    import os

    buf = BytesIO()
    w, h = A4
    c = canvas.Canvas(buf, pagesize=A4)
    LEFT  = 15 * mm
    RIGHT = w - 15 * mm

    GREEN_DARK = colors.HexColor("#1e3d2a")
    GREEN      = colors.HexColor("#2d7a4f")
    LGRAY      = colors.HexColor("#f2f2f2")
    DARK       = colors.HexColor("#1a1a2e")
    GRAY       = colors.HexColor("#6b7280")
    WHITE      = colors.white

    _ASSETS = os.path.join(os.path.dirname(__file__), "../assets")
    _LOGO   = os.path.join(_ASSETS, "LogoRecial.png")

    TOP = h - 15 * mm

    # ── Logo ─────────────────────────────────────────────────
    if os.path.exists(_LOGO):
        c.drawImage(_LOGO, RIGHT - 40*mm, TOP - 16*mm,
                    width=40*mm, height=16*mm,
                    preserveAspectRatio=True, mask='auto')

    # ── Title ─────────────────────────────────────────────────
    c.setFont("Helvetica-Bold", 16)
    c.setFillColor(GREEN_DARK)
    c.drawString(LEFT, TOP - 8*mm, "CUSTOMER LIST")

    c.setFont("Helvetica", 9)
    c.setFillColor(GRAY)
    c.drawString(LEFT, TOP - 14*mm,
        f"Generated: {date_type.today().strftime('%d/%m/%Y')}  ·  {len(customers)} customers")

    c.setStrokeColor(GREEN)
    c.setLineWidth(2)
    c.line(LEFT, TOP - 18*mm, RIGHT, TOP - 18*mm)

    table_y = TOP - 24*mm

    # ── Table ─────────────────────────────────────────────────
    headers = ["#", "Name", "CIF", "Address", "Email", "Phone"]
    data    = [headers]

    for i, cu in enumerate(customers, 1):
        data.append([
            str(i),
            cu.get("name") or "—",
            cu.get("cif")  or "—",
            cu.get("address") or "—",
            cu.get("email") or "—",
            cu.get("phone") or "—",
        ])

    col_widths = [10*mm, 42*mm, 27*mm, 47*mm, 36*mm, 22*mm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  GREEN_DARK),
        ("TEXTCOLOR",     (0,0), (-1,0),  WHITE),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0),  9),
        ("ALIGN",         (0,0), (-1,0),  "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, LGRAY]),
        ("TEXTCOLOR",     (0,1), (-1,-1), DARK),
        ("ALIGN",         (0,0), (0,-1),  "CENTER"),
        ("ALIGN",         (1,1), (-1,-1), "LEFT"),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#d1d5db")),
        ("ROWHEIGHT",     (0,0), (-1,-1), 8*mm),
        ("TOPPADDING",    (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
    ]))

    tbl_w, tbl_h = tbl.wrapOn(c, RIGHT - LEFT, h)
    tbl.drawOn(c, LEFT, table_y - tbl_h)

    # ── Footer ────────────────────────────────────────────────
    c.setFont("Helvetica", 8)
    c.setFillColor(GRAY)
    c.drawCentredString(w / 2, 10*mm,
        "RECICLAJES RECIAL S.L.  ·  C/ Carrera 56, 14880 Luque (Córdoba)  ·  info@recial.es")

    c.save()
    buf.seek(0)
    return buf.read()


def _generate_suppliers_pdf(suppliers: list) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from datetime import date as date_type
    import os

    buf = BytesIO()
    w, h = landscape(A4)
    c = canvas.Canvas(buf, pagesize=landscape(A4))
    LEFT  = 15 * mm
    RIGHT = w - 15 * mm

    GREEN_DARK  = colors.HexColor("#1e3d2a")
    GREEN       = colors.HexColor("#2d7a4f")
    GREEN_LIGHT = colors.HexColor("#8dc63f")
    LGRAY       = colors.HexColor("#f2f2f2")
    DARK        = colors.HexColor("#1a1a2e")
    GRAY        = colors.HexColor("#6b7280")
    WHITE       = colors.white
    BLUE        = colors.HexColor("#1d4ed8")

    _ASSETS = os.path.join(os.path.dirname(__file__), "../assets")
    _LOGO   = os.path.join(_ASSETS, "LogoRecial.png")

    TOP = h - 15 * mm

    horeca = [s for s in suppliers if s.get("supplier_type") == "Horeca"]
    urban  = [s for s in suppliers if s.get("supplier_type") == "Urban"]

    # ── Logo ─────────────────────────────────────────────────
    if os.path.exists(_LOGO):
        c.drawImage(_LOGO, RIGHT - 40*mm, TOP - 16*mm,
                    width=40*mm, height=16*mm,
                    preserveAspectRatio=True, mask='auto')

    # ── Title ─────────────────────────────────────────────────
    c.setFont("Helvetica-Bold", 16)
    c.setFillColor(GREEN_DARK)
    c.drawString(LEFT, TOP - 8*mm, "SUPPLIER LIST")

    c.setFont("Helvetica", 9)
    c.setFillColor(GRAY)
    c.drawString(LEFT, TOP - 14*mm,
        f"Generated: {date_type.today().strftime('%d/%m/%Y')}  ·  "
        f"{len(suppliers)} suppliers  ·  {len(horeca)} Horeca  ·  {len(urban)} Urban")

    c.setStrokeColor(GREEN)
    c.setLineWidth(2)
    c.line(LEFT, TOP - 18*mm, RIGHT, TOP - 18*mm)

    table_y = TOP - 24*mm

    # ── Main suppliers table ──────────────────────────────────
    headers = ["#", "Name", "Type", "CIF", "Address", "Email", "Phone", "Pickup Pts"]
    data    = [headers]

    for i, s in enumerate(suppliers, 1):
        pp_count = len(s.get("pickup_points", []))
        data.append([
            str(i),
            s.get("name") or "—",
            s.get("supplier_type") or "—",
            s.get("cif") or "—",
            s.get("address") or "—",
            s.get("email") or "—",
            s.get("phone") or "—",
            str(pp_count) if pp_count > 0 else "—",
        ])

    col_widths = [10*mm, 55*mm, 22*mm, 28*mm, 65*mm, 48*mm, 25*mm, 18*mm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  GREEN_DARK),
        ("TEXTCOLOR",     (0,0), (-1,0),  WHITE),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0),  9),
        ("ALIGN",         (0,0), (-1,0),  "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, LGRAY]),
        ("TEXTCOLOR",     (0,1), (-1,-1), DARK),
        ("ALIGN",         (0,0), (0,-1),  "CENTER"),
        ("ALIGN",         (2,0), (2,-1),  "CENTER"),
        ("ALIGN",         (7,0), (7,-1),  "CENTER"),
        ("ALIGN",         (1,1), (1,-1),  "LEFT"),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#d1d5db")),
        ("ROWHEIGHT",     (0,0), (-1,-1), 8*mm),
        ("TOPPADDING",    (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
    ]))

    # Colour Type column per row
    for i, s in enumerate(suppliers, 1):
        is_horeca = s.get("supplier_type") == "Horeca"
        bg = colors.HexColor("#eff6ff") if is_horeca else colors.HexColor("#f0fdf4")
        fc = BLUE if is_horeca else GREEN
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (2,i), (2,i), bg),
            ("TEXTCOLOR",  (2,i), (2,i), fc),
            ("FONTNAME",   (2,i), (2,i), "Helvetica-Bold"),
        ]))

    tbl_w, tbl_h = tbl.wrapOn(c, RIGHT - LEFT, h)
    tbl.drawOn(c, LEFT, table_y - tbl_h)

    # ── Urban pickup points section ───────────────────────────
    pickup_y = table_y - tbl_h - 14*mm
    urban_with_points = [s for s in urban if s.get("pickup_points")]

    if urban_with_points:
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(GREEN_DARK)
        c.drawString(LEFT, pickup_y, "URBAN SUPPLIERS — PICKUP POINTS WITH COORDINATES")

        c.setStrokeColor(GREEN_LIGHT)
        c.setLineWidth(1.5)
        c.line(LEFT, pickup_y - 3*mm, RIGHT, pickup_y - 3*mm)
        pickup_y -= 10*mm

        for s in urban_with_points:
            pps = s.get("pickup_points", [])

            # Supplier sub-header
            c.setFont("Helvetica-Bold", 10)
            c.setFillColor(GREEN)
            c.drawString(LEFT, pickup_y, f"  {s['name']}")
            pickup_y -= 6*mm

            pp_headers = ["Pickup Point Name", "Latitude", "Longitude", "Notes"]
            pp_data    = [pp_headers]
            for pp in pps:
                lat = f"{pp['latitude']:.6f}"  if pp.get("latitude")  else "—"
                lng = f"{pp['longitude']:.6f}" if pp.get("longitude") else "—"
                pp_data.append([
                    pp.get("name") or "—",
                    lat,
                    lng,
                    pp.get("notes") or "",
                ])

            pp_col_widths = [80*mm, 38*mm, 38*mm, 115*mm]
            pp_tbl = Table(pp_data, colWidths=pp_col_widths)
            pp_tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,0),  GREEN_LIGHT),
                ("TEXTCOLOR",     (0,0), (-1,0),  WHITE),
                ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
                ("FONTSIZE",      (0,0), (-1,0),  8),
                ("ALIGN",         (0,0), (-1,0),  "CENTER"),
                ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
                ("FONTSIZE",      (0,1), (-1,-1), 8),
                ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, LGRAY]),
                ("TEXTCOLOR",     (0,1), (-1,-1), DARK),
                ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                ("ALIGN",         (1,1), (2,-1),  "CENTER"),
                ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#d1d5db")),
                ("ROWHEIGHT",     (0,0), (-1,-1), 7*mm),
                ("LEFTPADDING",   (0,0), (-1,-1), 4),
                ("TOPPADDING",    (0,0), (-1,-1), 2),
                ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ]))

            pp_tbl_w, pp_tbl_h = pp_tbl.wrapOn(c, RIGHT - LEFT, h)
            pp_tbl.drawOn(c, LEFT, pickup_y - pp_tbl_h)
            pickup_y -= pp_tbl_h + 8*mm

    # ── Footer ────────────────────────────────────────────────
    c.setFont("Helvetica", 8)
    c.setFillColor(GRAY)
    c.drawCentredString(w / 2, 8*mm,
        "RECICLAJES RECIAL S.L.  ·  C/ Carrera 56, 14880 Luque (Córdoba)  ·  info@recial.es")

    c.save()
    buf.seek(0)
    return buf.read()


@router.get("/customers-list")
def get_customers_list(db: Session = Depends(get_db)):
    from models.customers import Customer
    customers = db.query(Customer).order_by(Customer.name).all()
    return {
        "total": len(customers),
        "customers": [{
            "id":      c.id,
            "name":    c.name,
            "cif":     getattr(c, "cif", None),
            "address": getattr(c, "address", None),
            "email":   getattr(c, "email", None),
            "phone":   getattr(c, "phone", None),
        } for c in customers]
    }


@router.get("/customers-list/pdf")
def get_customers_list_pdf(db: Session = Depends(get_db)):
    from models.customers import Customer
    customers = db.query(Customer).order_by(Customer.name).all()
    data = [{
        "name":    c.name,
        "cif":     getattr(c, "cif", None),
        "address": getattr(c, "address", None),
        "email":   getattr(c, "email", None),
        "phone":   getattr(c, "phone", None),
    } for c in customers]
    pdf = _generate_customers_pdf(data)
    return StreamingResponse(
        BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=Recial_Customers.pdf"}
    )