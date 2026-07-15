from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile
from sqlalchemy.orm import Session
from typing import Optional
from openpyxl import load_workbook
from io import BytesIO

from database import get_db
from models.suppliers import Supplier, SupplierType
from models.receipts import Receipt
from schemas.suppliers import (
    SupplierCreate,
    SupplierUpdate,
    SupplierResponse,
    SupplierListResponse,
)

from auth import get_current_user, require_admin, require_manager_or_above
from models.users import User
from models.pickupPoints import PickupPoint 

router = APIRouter(prefix="/suppliers", tags=["Suppliers"])

# ── Supplier import helpers ─────────────────────────────────
_COLUMN_MAP = {
    "contact person": "contact_person",
    "name":           "name",
    "cif":            "cif",
    "address":        "address",
    "town":           "city",       # sheet "Town"   -> model city
    "county":         "county",     # sheet "County" -> model county
    "phone":          "phone",
    "email":          "email",
}

# ── Supplier import helpers ─────────────────────────────────

def _clean(v):
    """Trim strings; blanks -> None. Strip trailing .0 from Excel numerics."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return s or None


def _to_float(v):
    """
    Parse a coordinate. Handles real floats, strings, and Spanish comma
    decimals ("-4,63600" -> -4.636). Returns None if empty/unparseable.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


_COUNTY_FIXES = {"CORODBA": "CORDOBA"}


def _fix_county(v):
    if not v:
        return None
    return _COUNTY_FIXES.get(v.upper(), v)


def _header_map(header_row):
    """Map lowercased header names -> column index."""
    idx = {}
    for i, h in enumerate(header_row):
        k = _clean(h)
        if k:
            idx[k.lower()] = i
    return idx


def _get(row, hmap, key):
    """Fetch a cell by header name (case-insensitive)."""
    i = hmap.get(key.lower())
    if i is None or i >= len(row):
        return None
    return row[i]

@router.get("/", response_model=SupplierListResponse)
def get_suppliers(
    skip: int = 0,
    limit: int = 50,
    search: Optional[str] = None,
    supplier_type: Optional[SupplierType] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Supplier)
    if search:
        query = query.filter(
            Supplier.name.ilike(f"%{search}%") |
            Supplier.cif.ilike(f"%{search}%") |
            Supplier.address.ilike(f"%{search}%")
        )
    if supplier_type:
        query = query.filter(Supplier.supplier_type == supplier_type)
    total = query.count()
    suppliers = query.order_by(Supplier.name).offset(skip).limit(limit).all()
    return {"total": total, "suppliers": suppliers}


@router.get("/{supplier_id}", response_model=SupplierResponse)
def get_supplier(supplier_id: int, db: Session = Depends(get_db)):
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return supplier


@router.post("/", response_model=SupplierResponse, status_code=201)
def create_supplier(supplier_data: SupplierCreate, db: Session = Depends(get_db), current_user: User = Depends(require_manager_or_above) ):
    if supplier_data.cif:
        existing = db.query(Supplier).filter(
            Supplier.cif == supplier_data.cif
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="CIF already registered")
    new_supplier = Supplier(**supplier_data.model_dump())
    db.add(new_supplier)
    db.commit()
    db.refresh(new_supplier)
    return new_supplier


@router.patch("/{supplier_id}", response_model=SupplierResponse)
def update_supplier(
    supplier_id: int,
    supplier_data: SupplierUpdate,
    db: Session = Depends(get_db),
):
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    for field, value in supplier_data.model_dump(exclude_unset=True).items():
        setattr(supplier, field, value)
    db.commit()
    db.refresh(supplier)
    return supplier


@router.delete("/{supplier_id}", status_code=204)
def delete_supplier(supplier_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    receipt_count = db.query(Receipt).filter(Receipt.supplier_id == supplier_id).count()
    if receipt_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"No se puede eliminar este proveedor porque tiene {receipt_count} albarán(es) registrado(s). "
                f"Los albaranes son registros de trazabilidad y deben conservarse."
        )

    db.delete(supplier)
    db.commit()
    return None



@router.post("/import")
async def import_suppliers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx).")
 
    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400,
            detail=f"No se pudo leer el archivo Excel: {type(e).__name__}")
 
    # Existing supplier names (dedupe key — CIF is NOT unique in this data)
    existing_names = {
        (n or "").strip().lower() for (n,) in db.query(Supplier.name).all()
    }
 
    created_horeca, created_urban, skipped = [], [], []
    total_pickup_points = 0
    seen = set()
 
    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1 — HORECA (flat; coordinates on the supplier)
    # ══════════════════════════════════════════════════════════════════════
    if "Horeca" in wb.sheetnames:
        ws = wb["Horeca"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header:
            hmap = _header_map(header)
            for row in rows:
                name = _clean(_get(row, hmap, "name"))
                if not name:
                    continue
                key = name.lower()
                if key in existing_names:
                    skipped.append({"name": name, "reason": "ya existe"}); continue
                if key in seen:
                    skipped.append({"name": name, "reason": "duplicado en el archivo"}); continue
                seen.add(key)
 
                db.add(Supplier(
                    supplier_type=SupplierType.HORECA,
                    name=name,
                    cif=_clean(_get(row, hmap, "cif")),
                    address=_clean(_get(row, hmap, "address")),
                    city=_clean(_get(row, hmap, "Town")),
                    county=_fix_county(_clean(_get(row, hmap, "County"))),
                    contact_person=_clean(_get(row, hmap, "Contact Person")),
                    phone=_clean(_get(row, hmap, "Phone")),
                    email=_clean(_get(row, hmap, "EMAIL")),
                    latitude=_to_float(_get(row, hmap, "Latitude")),
                    longitude=_to_float(_get(row, hmap, "Longitude")),
                    is_active=True,
                ))
                created_horeca.append(name)
 
    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2 — URBAN (parent supplier row, then its pickup-point rows)
    # ══════════════════════════════════════════════════════════════════════
    if "Urban" in wb.sheetnames:
        ws = wb["Urban"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header:
            hmap = _header_map(header)
            current = None          # the Supplier object we're attaching points to
 
            for row in rows:
                name = _clean(_get(row, hmap, "name"))
                pp_name = _clean(_get(row, hmap, "Pickup point"))
 
                # ── Parent row: a new council/supplier ──
                if name:
                    key = name.lower()
                    if key in existing_names or key in seen:
                        reason = "ya existe" if key in existing_names else "duplicado en el archivo"
                        skipped.append({"name": name, "reason": reason})
                        current = None          # skip its pickup points too
                        continue
                    seen.add(key)
 
                    current = Supplier(
                        supplier_type=SupplierType.URBAN,
                        name=name,
                        cif=_clean(_get(row, hmap, "cif")),
                        address=_clean(_get(row, hmap, "Address")),
                        city=_clean(_get(row, hmap, "Town")),
                        county=_fix_county(_clean(_get(row, hmap, "County"))),
                        contact_person=_clean(_get(row, hmap, "Contact Person")),
                        is_active=True,
                    )
                    db.add(current)
                    created_urban.append(name)
                    continue
 
                # ── Child row: a pickup point for the current supplier ──
                if pp_name and current is not None:
                    current.pickup_points.append(PickupPoint(
                        name=pp_name,
                        address=_clean(_get(row, hmap, "Address")),
                        latitude=_to_float(_get(row, hmap, "Latitude")),
                        longitude=_to_float(_get(row, hmap, "Longitude")),
                    ))
                    total_pickup_points += 1
 
    db.commit()
 
    return {
        "created_count":  len(created_horeca) + len(created_urban),
        "horeca_count":   len(created_horeca),
        "urban_count":    len(created_urban),
        "pickup_points":  total_pickup_points,
        "skipped_count":  len(skipped),
        "skipped":        skipped,
    }
 





















































































































